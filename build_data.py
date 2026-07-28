#!/usr/bin/env python3
"""
Build data.json for the Hajj & Umrah open-data dashboard.

Source: Ministry of Hajj & Umrah open data (https://haj.gov.sa/en/Open-Data)

TRUNCATION RULE (enforced here, not left to the reader):
  All 1447 files were last published 2026-03-05, mid-Ramadan 1447. Ramadan is the
  final month present and is a PARTIAL month. Any YoY figure that includes it is
  wrong. Partial months are flagged `partial: true` and EXCLUDED from every
  aggregate and growth calculation below.

PRIVACY: only aggregates are emitted. The agent-level files contain personal
email addresses and are never read here.
"""
import json, os, statistics, sys

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

F = os.environ.get("MOHU_FILES", "./source-data/files")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

MONTHS = ["Muharram", "Safar", "Rabi I", "Rabi II", "Jumada I", "Jumada II",
          "Rajab", "Sha'ban", "Ramadan", "Shawwal", "Dhu al-Qa'dah", "Dhu al-Hijjah"]


def rows(name):
    p = os.path.join(F, name + ".xlsx")
    if not os.path.exists(p):
        print("  ! missing:", name)
        return []
    ws = openpyxl.load_workbook(p, read_only=True, data_only=True).worksheets[0]
    return [r for r in ws.iter_rows(values_only=True) if r and any(c is not None for c in r)]


def num(x):
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return None


def month_series(name, year):
    """Rows are chronological. Return ordered values for the target year."""
    out = []
    for r in rows(name)[1:]:
        if r[0] is None or str(r[0]).strip() != year:
            continue
        v = num(r[2] if len(r) > 2 else None)
        if v is None:
            continue
        out.append(int(v))
    return out


D = {"meta": {
    "source": "Ministry of Hajj & Umrah Open Data",
    "source_url": "https://haj.gov.sa/en/Open-Data",
    "licence": "Saudi Open Data License — reuse permitted with attribution",
    "captured": "2026-07-28",
    "truncation_rule": ("1447 files were published 2026-03-05, mid-Ramadan 1447. "
                        "Ramadan 1447 is a partial month and is excluded from all "
                        "totals and growth figures on this page."),
}}

# ---------- 1. visas by season ----------
season = []
for r in rows("Number-of-Umrah-performers-by-year")[1:]:
    s, v = str(r[0]).strip(), num(r[1])
    if v:
        season.append({"season": s, "visas": int(v), "covid": s in ("1441", "1442", "1443")})
season.sort(key=lambda x: x["season"])
D["visas_by_season"] = season

# ---------- 2. monthly, 1446 vs 1447 ----------
e46 = month_series("Number-of-Umrah-entrants-for-the-year-1446-by-month", "1446")
e47 = month_series("Number-of-Umrah-Entrants-in-1447-AH-by-Month", "1447")
v46 = month_series("Number-of-Umrah-visas-for-the-year-1446-by-month", "1446")
v47 = month_series("Number-of-Umrah-Visas-by-Month-for-the-Year-1447-AH", "1447")

def drop_junk_tail(series):
    """Trailing months with near-zero counts are reporting artifacts, not data.
    (1446 entrants ends with a month of `2`.) Drop them."""
    if not series:
        return series
    med = statistics.median(series)
    while series and series[-1] < med * 0.01:
        series = series[:-1]
    return series

e46, v46 = drop_junk_tail(e46), drop_junk_tail(v46)

n47 = len(e47)
complete = n47 - 1                      # last 1447 month is partial
monthly = []
for i in range(max(len(e46), n47)):
    monthly.append({
        "month": MONTHS[i] if i < len(MONTHS) else f"M{i+1}",
        "e1446": e46[i] if i < len(e46) else None,
        "e1447": e47[i] if i < n47 else None,
        "v1446": v46[i] if i < len(v46) else None,
        "v1447": v47[i] if i < len(v47) else None,
        "partial": i == n47 - 1,
    })
D["monthly"] = monthly

yoy = []
for i in range(complete):
    if i < len(e46) and i < len(e47) and e46[i]:
        yoy.append({"month": MONTHS[i],
                    "entrants": round((e47[i] - e46[i]) / e46[i] * 100, 1),
                    "visas": round((v47[i] - v46[i]) / v46[i] * 100, 1) if i < len(v46) and v46[i] else None})
D["yoy_by_month"] = yoy

E46, E47 = sum(e46[:complete]), sum(e47[:complete])
V46, V47 = sum(v46[:complete]), sum(v47[:complete])
D["growth"] = {
    "complete_months": complete,
    "entrants_1446": E46, "entrants_1447": E47,
    "entrants_pct": round((E47 - E46) / E46 * 100, 1),
    "visas_1446": V46, "visas_1447": V47,
    "visas_pct": round((V47 - V46) / V46 * 100, 1),
    "full_year_1446_visas": season[-1]["visas"] if season else None,
}

# ---------- 3. length of stay ----------
stay = []
for r in rows("Average-stay-of-Umrah-performers-by-season")[1:]:
    s, v = str(r[0]).strip(), num(r[1])
    if v:
        stay.append({"season": s, "nights": round(v, 2), "covid": s in ("1441", "1442", "1443")})
stay.sort(key=lambda x: x["season"])
D["stay_by_season"] = stay

# ---------- 4. accommodation class & continent ----------
cls, cont = {}, {}
for r in rows("Number-of-Umrah-Performers-based-on-Hotels-and-Continents-Classification"):
    c = [x for x in r if x is not None]
    if len(c) < 3:
        continue
    v = num(c[2])
    if v is None:
        continue
    cls[str(c[0]).strip()] = cls.get(str(c[0]).strip(), 0) + v
    cont[str(c[1]).strip()] = cont.get(str(c[1]).strip(), 0) + v
EN = {"حد ادنى": "Minimum grade", "حد أدنى": "Minimum grade", "3 نجوم": "3-star",
      "نجمتين": "2-star", "4 نجوم": "4-star", "5 نجوم": "5-star",
      "الدرجة الثالثة": "Third grade", "الدرجة الثانية": "Second grade",
      "اسيا": "Asia", "أفريقيا": "Africa", "أوربا": "Europe",
      "أمريكا الشمالية": "North America", "أمريكا الجنوبية": "South America",
      "أقيانوسيا- استراليا": "Oceania"}
ORDER = ["5-star", "4-star", "3-star", "2-star", "Minimum grade", "Third grade", "Second grade"]
D["hotel_class"] = sorted(
    [{"class": EN.get(k, k), "pilgrims": int(v)} for k, v in cls.items() if v],
    key=lambda x: ORDER.index(x["class"]) if x["class"] in ORDER else 99)
D["continent"] = sorted(
    [{"continent": EN.get(k, k), "pilgrims": int(v)} for k, v in cont.items() if v],
    key=lambda x: -x["pilgrims"])

# ---------- 5. agents by country ----------
def agents(name):
    d = {}
    for r in rows(name)[1:]:
        c = [x for x in r if x is not None]
        ctry = next((str(x).strip() for x in c
                     if isinstance(x, str) and not str(x).strip().replace(".", "").isdigit()), None)
        v = next((num(x) for x in reversed(c) if num(x) is not None), None)
        if ctry and v is not None:
            d[ctry] = int(v)
    return d

a46 = agents("Number-of-active-agents-for-the-year-1446-by-country")
a47 = agents("Number-of-Active-Agents-for-the-Year-1447-AH-by-Country")
AR2EN = {"مصر": "Egypt", "الجزائر": "Algeria", "باكستان": "Pakistan", "إندونيسيا": "Indonesia",
         "اندونيسيا": "Indonesia", "المغرب": "Morocco", "السودان": "Sudan", "الهند": "India",
         "اليمن": "Yemen", "بنجلادش": "Bangladesh", "بنغلاديش": "Bangladesh", "تركيا": "Turkey",
         "نيجيريا": "Nigeria", "سوريا": "Syria", "العراق": "Iraq", "أفغانستان": "Afghanistan",
         "فرنسا": "France", "الأردن": "Jordan", "ماليزيا": "Malaysia", "ألمانيا": "Germany",
         "الولايات المتحدة الأمريكية": "United States", "فلسطين": "Palestine", "تونس": "Tunisia",
         "ليبيا": "Libya", "الإمارات العربية المتحدة": "UAE", "بريطانيا": "United Kingdom"}
a46en = {AR2EN.get(k, k): v for k, v in a46.items()}
top = sorted(a47.items(), key=lambda x: -x[1])[:14]
D["agents"] = {
    "total_1446": sum(a46.values()), "total_1447": sum(a47.values()),
    "countries_1446": len(a46), "countries_1447": len(a47),
    "top": [{"country": k, "y1447": v, "y1446": a46en.get(k)} for k, v in top],
}

# ---------- 6. permit composition ----------
perm = {}
for r in rows("Umrah-permits-for-the-year-1446-by-month-and-type-1")[1:]:
    c = [x for x in r if x is not None]
    v = next((num(x) for x in reversed(c) if num(x) is not None), None)
    t = None
    for x in c:
        if isinstance(x, str) and not str(x).strip().isdigit() and len(str(x).strip()) > 2:
            t = str(x).strip()
    if t and v:
        perm[t] = perm.get(t, 0) + v
PEN = {"معتمر خارجي": "External Umrah visa", "مواطن": "Saudi citizen", "مقيم": "Resident",
       "زائر": "Visit-visa holder", "خليجي": "GCC national"}
D["permit_types"] = sorted(
    [{"type": PEN.get(k, k), "permits": int(v)} for k, v in perm.items() if v],
    key=lambda x: -x["permits"])

# ---------- 7. entry channel ----------
ports = {}
for r in rows("Number-of-Umrah-Entrants-in-1447-AH-by-Port-Type")[1:]:
    c = [x for x in r if x is not None]
    lab = next((str(x).strip() for x in c if isinstance(x, str) and not str(x).strip().isdigit()), None)
    v = next((num(x) for x in reversed(c) if num(x) is not None), None)
    if lab and v:
        ports[lab] = ports.get(lab, 0) + v
D["entry_channel"] = sorted([{"channel": k, "entrants": int(v)} for k, v in ports.items()],
                            key=lambda x: -x["entrants"])

# ---------- 8. Rawdah scarcity (Ramadan 1446) ----------
def daily(name):
    out = []
    for r in rows(name)[1:]:
        v = next((num(x) for x in reversed([y for y in r if y is not None]) if num(x) is not None), None)
        if v:
            out.append(v)
    return out

rd, um = daily("Number-of-Rawdah-permits-in-days-for-Ramadan-1446"), daily("Umrah-permits-in-days-for-Ramadan-1446")
if rd and um:
    D["rawdah"] = {
        "rawdah_mean": round(statistics.mean(rd)), "rawdah_max": int(max(rd)),
        "rawdah_sd": round(statistics.pstdev(rd)),
        "umrah_mean": round(statistics.mean(um)),
        "ratio": round(statistics.mean(um) / statistics.mean(rd), 1),
    }

# ---------- 9. domestic Hajj ----------
dom = rows("Number-of-domestic-pilgrims-by-company-package-and-city-for-the-year-1446")
if dom:
    tot, comps = 0, set()
    for r in dom[1:]:
        v = next((num(x) for x in reversed([y for y in r if y is not None]) if num(x) is not None), None)
        s = [str(x).strip() for x in r if isinstance(x, str)]
        if v:
            tot += v
            if s: comps.add(s[0])
    D["domestic_hajj"] = {"pilgrims_1446": int(tot), "companies": len(comps)}

# =====================================================================
#  TOURISM — Ministry of Tourism, via the national open data platform
#  Same truncation discipline: 2025 is a PART year (Jan–Jun) and is
#  flagged, never summed against a complete year.
# =====================================================================
import glob

TF = os.environ.get("TOURISM_FILES", "./source-data/tourism")

def trows(pat):
    g = sorted(glob.glob(os.path.join(TF, pat + "*.xlsx")))
    if not g:
        print("  ! tourism missing:", pat)
        return []
    ws = openpyxl.load_workbook(g[0], read_only=True, data_only=True).worksheets[0]
    return [r for r in ws.iter_rows(values_only=True) if r and any(c is not None for c in r)]

SKIP_TOTAL = ("الإجمالي", "اجمالي", "TOTAL", "Grand Total", "إجمالي المملكة")
MON = ["January", "February", "March", "April", "May", "June",
       "July", "August", "September", "October", "November", "December"]

def flow(pat):
    """monthly series: tourists (000), overnight (000), spend (SAR mn), LOS, spend/trip"""
    out = []
    for r in trows(pat)[1:]:
        if r[0] is None or r[1] is None:
            continue
        m = str(r[1]).strip()
        if m not in MON:            # skips "Grand Total"
            continue
        out.append({"year": str(r[0]).strip(), "month": m,
                    "tourists": num(r[2]), "nights": num(r[3]), "spend": num(r[4]),
                    "los": num(r[5]), "per_trip": num(r[6])})
    return out

inb, dom = flow("Inbound-Tourism-Statistics-H1-2025"), flow("Domestic-Tourism-Statistics-H1-2025")

def annual(series):
    y = {}
    for r in series:
        a = y.setdefault(r["year"], {"tourists": 0, "nights": 0, "spend": 0, "months": 0})
        a["tourists"] += r["tourists"] or 0
        a["nights"] += r["nights"] or 0
        a["spend"] += r["spend"] or 0
        a["months"] += 1
    return [{"year": k, "tourists": round(v["tourists"]), "nights": round(v["nights"]),
             "spend": round(v["spend"]), "months": v["months"], "partial": v["months"] < 12}
            for k, v in sorted(y.items())]

T = {"annual_inbound": annual(inb), "annual_domestic": annual(dom)}

# monthly, aligned for charting
bykey = {}
for r in inb: bykey[(r["year"], r["month"])] = {"inbound": r}
for r in dom: bykey.setdefault((r["year"], r["month"]), {})["domestic"] = r
T["monthly"] = [{"year": k[0], "month": k[1],
                 "inbound": round(v.get("inbound", {}).get("tourists") or 0),
                 "domestic": round(v.get("domestic", {}).get("tourists") or 0),
                 "inbound_los": round(v.get("inbound", {}).get("los") or 0, 2),
                 "domestic_los": round(v.get("domestic", {}).get("los") or 0, 2)}
                for k, v in sorted(bykey.items(), key=lambda x: (x[0][0], MON.index(x[0][1])))]

# region name map (the accommodation file carries both Arabic and English)
AR_REG = {"منطقة المدينة المنورة":"Madinah","منطقة مكة المكرمة":"Makkah","منطقة الرياض":"Riyadh",
          "المنطقة الشرقية":"Eastern","منطقة عسير":"Aseer","منطقة القصيم":"Qassim",
          "منطقة تبوك":"Tabuk","منطقة حائل":"Hail","منطقة جازان":"Jazan","منطقة نجران":"Najran",
          "منطقة الباحة":"Albaha","منطقة الجوف":"Jouf","منطقة الحدود الشمالية":"Northern Borders"}
regmap, fac = dict(AR_REG), []
for r in trows("Accommodation-per-destination-2024"):
    c = list(r) + [None] * 6
    ar, ap, ho, tt, en = c[1], c[2], c[3], c[4], c[5]
    if ar and en and num(ap) is not None and num(ho) is not None:
        ar = str(ar).strip(); en = str(en).strip()
        regmap[ar] = en
        if en.upper() not in ("TOTAL", "GRAND TOTAL") and ar not in SKIP_TOTAL:
            fac.append({"province": en, "apartments": int(num(ap)), "hotels": int(num(ho)),
                        "total": int(num(tt) or 0)})
T["facilities"] = sorted(fac, key=lambda x: -x["total"])

# occupancy by region & facility type (2024 average)
occ = {}
for r in trows("Apartment-occupancy-rates-by-region-2024")[1:]:
    if r[1] is None:
        continue
    reg = str(r[1]).strip(); typ = str(r[2]).strip() if r[2] else ""
    vals = [num(x) for x in r[3:15] if num(x) is not None]
    if not vals:
        continue
    avg = sum(vals) / len(vals)
    if reg in SKIP_TOTAL: continue
    e = occ.setdefault(regmap.get(reg, reg), {})
    e["hotels" if "فناد" in typ else "apartments"] = round(avg * 100, 1)
T["occupancy"] = sorted([{"region": k, **v} for k, v in occ.items() if v.get("hotels")],
                        key=lambda x: -(x.get("hotels") or 0))

# spending by purpose
PUR = {"الترفيه": "Leisure", "عمل": "Business",
       "زيارة الأصدقاء أو الأقارب": "Visiting friends / family",
       "أخرى": "Other", "التعليم والتدريب": "Education", "الصحة": "Health",
       "غرض ديني": "Religious", "ديني": "Religious", "التسوق": "Shopping"}
CATS = ["Accommodation", "Entertainment", "Food & beverage", "Transport"]
def purpose(pat):
    out = []
    for r in trows(pat)[1:]:
        if r[2] is None:
            continue
        vals = [num(x) or 0 for x in r[3:7]]
        if str(r[2]).strip() in SKIP_TOTAL: continue
        out.append({"purpose": PUR.get(str(r[2]).strip(), str(r[2]).strip()),
                    "values": [round(v) for v in vals], "total": round(sum(vals))})
    return sorted(out, key=lambda x: -x["total"])
T["spend_categories"] = CATS
T["spend_inbound"] = purpose("Distribution-of-inbound-tourism-spending-by-purpose-2025")
T["spend_domestic"] = purpose("Distribution-of-spending-on-domestic-tourism-by-purpose-2025")
T["meta"] = {"source": "Ministry of Tourism, via the Saudi National Open Data Platform",
             "source_url": "https://open.data.gov.sa/en/datasets?category=tourism",
             "note": "2025 is a part year (January–June) and is flagged wherever it appears."}
D["tourism"] = T

json.dump(D, open(OUT, "w"), ensure_ascii=False, separators=(",", ":"))
# data.js as well, so the page renders from file:// without a server
with open(os.path.join(os.path.dirname(OUT), "data.js"), "w") as fh:
    fh.write("window.DATA=" + json.dumps(D, ensure_ascii=False, separators=(",", ":")) + ";")
kb = round(os.path.getsize(OUT) / 1024, 1)
print(f"wrote {OUT}  ({kb} KB)")
print(f"  seasons={len(D['visas_by_season'])} monthly={len(D['monthly'])} "
      f"yoy={len(D['yoy_by_month'])} (complete months only)")
print(f"  growth: entrants {D['growth']['entrants_pct']:+}%  visas {D['growth']['visas_pct']:+}%")
print(f"  agents {D['agents']['total_1446']:,} -> {D['agents']['total_1447']:,}")
